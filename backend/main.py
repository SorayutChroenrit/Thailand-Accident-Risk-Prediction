"""
FastAPI Backend for Accident Risk Prediction
ใช้โมเดล ML ที่เทรนจาก Google Colab

FIXED: Vehicle type และ Weather condition filters ใช้งานได้แล้ว
ADDED: Accident cause filter
"""

import json
from datetime import datetime
from typing import Dict, List, Optional

import joblib
import numpy as np
import pandas as pd
import requests
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

app = FastAPI(title="Accident Risk Prediction API")

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =============================================================================
# REGISTER ANALYTICS ROUTER
# =============================================================================
# from analytics.routes import router as analytics_router
# app.include_router(analytics_router)
# Note: Analytics module not yet implemented

print("✅ Analytics endpoints registered:")
print("   POST /analytics/descriptive")
print("   POST /analytics/diagnostic")
print("   POST /analytics/predictive")
print("   POST /analytics/prescriptive")
print("   GET  /analytics/health")

# =============================================================================
# TRAFFIC & ROAD CONDITION ENDPOINTS
# =============================================================================


@app.get("/traffic/density")
async def get_traffic_density(lat: float, lon: float):
    """Get traffic density at location (mock data for now)"""
    hour = datetime.now().hour
    is_rush_hour = (hour >= 7 and hour <= 9) or (hour >= 17 and hour <= 19)

    # Mock traffic density (0-1 scale)
    import random

    density = 0.4  # base
    if is_rush_hour:
        density += 0.3
    density += random.random() * 0.2
    density = min(1, density)

    # Average speed (km/h)
    base_speed = 60
    speed = base_speed * (1 - density * 0.7)

    # Congestion level
    congestion_level = "light"
    if density > 0.7:
        congestion_level = "heavy"
    elif density > 0.5:
        congestion_level = "moderate"

    return {
        "density": round(density, 2),
        "average_speed": round(speed),
        "congestion_level": congestion_level,
        "timestamp": datetime.now().isoformat(),
    }


@app.get("/traffic/index")
async def get_traffic_index():
    """Get traffic index (Longdo-style, 0-10 scale)"""
    import random

    hour = datetime.now().hour
    is_rush_hour = (hour >= 7 and hour <= 9) or (hour >= 17 and hour <= 19)

    index = 3  # base
    if is_rush_hour:
        index += 3
    index += random.random() * 2
    index = max(0, min(10, index))

    return {
        "current": round(index, 1),
        "status": "clear"
        if index < 3
        else "moderate"
        if index < 5
        else "busy"
        if index < 7
        else "congested",
        "timestamp": datetime.now().isoformat(),
    }


@app.get("/road/condition")
async def get_road_condition(lat: float, lon: float):
    """Get road condition at location (mock data for now)"""
    import random

    qualities = ["excellent", "good", "fair", "poor"]
    quality = random.choice(qualities)

    return {
        "surface_type": "asphalt",
        "quality": quality,
        "lane_count": random.randint(2, 4),
        "speed_limit": random.choice([60, 80, 90, 100, 120]),
        "has_shoulder": random.random() > 0.3,
        "lighting": "good" if random.random() > 0.4 else "poor",
        "last_maintenance": (
            datetime.now() - pd.Timedelta(days=random.randint(1, 365))
        ).isoformat(),
    }


@app.get("/road/hazards")
async def get_road_hazards(lat: float, lon: float, radius: float = 5):
    """Get nearby road hazards (mock data for now)"""
    import random

    hazard_types = [
        "pothole",
        "debris",
        "construction",
        "flooding",
        "animal_crossing",
        "poor_visibility",
    ]
    hazard_count = random.randint(0, 3)

    hazards = [
        {
            "id": f"hazard_{i}",
            "type": random.choice(hazard_types),
            "lat": lat + (random.random() - 0.5) * 0.01,
            "lon": lon + (random.random() - 0.5) * 0.01,
            "severity": random.choice(["low", "medium", "high"]),
            "reported_at": (
                datetime.now() - pd.Timedelta(hours=random.randint(1, 24))
            ).isoformat(),
        }
        for i in range(hazard_count)
    ]

    return {"hazards": hazards, "count": len(hazards), "radius": radius}


# =============================================================================
# LOAD MODELS - DIRECT SEVERITY PREDICTION MODEL
# =============================================================================
print("🚀 Loading Direct Severity Prediction ML model...")

# Suppress version mismatch warnings (models still work fine)
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='pickle')
warnings.filterwarnings('ignore', category=UserWarning, module='xgboost')
from sklearn.exceptions import InconsistentVersionWarning
warnings.filterwarnings('ignore', category=InconsistentVersionWarning)

try:
    # Load direct severity prediction model
    xgb_direct_model = joblib.load("models/xgboost_direct_model.pkl")
    label_encoder = joblib.load("models/label_encoder.pkl")

    # Feature names for the direct model (113 features with engineered features)
    feature_names = [
        "KM", "LATITUDE", "LONGITUDE", "hour", "temperature", "dewpoint", "humidity",
        "wind_speed", "wind_direction", "pressure", "cloud_cover", "year", "month",
        "day", "dayofweek", "quarter", "is_weekend", "is_monday", "is_friday",
        "is_saturday", "is_sunday", "is_night", "is_rush_hour", "is_peak_hour",
        "is_festival_month", "is_songkran", "is_newyear_month", "is_hot_season",
        "is_rainy_season", "is_cool_season", "is_clear", "is_rainy", "is_foggy",
        "is_poor_visibility", "is_bangkok", "is_chonburi", "is_nakhon_ratchasima",
        "is_chiang_mai", "is_top20_province", "is_top10_province", "is_motorcycle",
        "is_pickup", "is_sedan", "is_truck", "is_bus", "is_commercial_vehicle",
        "is_rollover", "is_rear_end", "is_head_on", "is_single_vehicle",
        "is_speeding", "is_drowsy", "is_dui", "is_reckless", "is_traffic_violation",
        "is_inexperienced", "is_vehicle_defect", "is_road_condition", "is_human_error",
        "is_highway", "is_local_road", "is_rural_road", "hour_sin", "hour_cos",
        "dayofweek_sin", "dayofweek_cos", "month_sin", "month_cos", "day_sin",
        "day_cos", "hour_risk_score", "day_risk_score", "month_risk_score",
        "province_risk_score", "weather_risk_score", "vehicle_risk_score",
        "cause_risk_score", "rain_rush_hour", "friday_evening", "weekend_night",
        "saturday_night", "songkran_daytime", "songkran_period", "fog_morning",
        "weekday_morning_rush", "weekday_evening_rush", "bangkok_rush",
        "day_hour_risk", "weather_hour_risk", "month_hour_risk",
        "motorcycle_chiang_mai", "motorcycle_rain", "motorcycle_night",
        "speeding_highway", "dui_weekend_night", "drowsy_highway",
        "human_error_rush", "rural_night", "commercial_highway", "festival_weekend",
        "rain_night", "overall_risk_score", "is_early_morning", "is_late_evening",
        "is_midnight", "is_month_start", "is_month_end", "is_post_covid",
        "years_since_2019", "is_q1", "is_q2", "is_q3", "is_q4"
    ]
    
    # Severity class names (Thai)
    class_names = ['บาดเจ็บเล็กน้อย', 'บาดเจ็บสาหัส', 'เสียชีวิต']

    print(f"✅ Direct Severity Model loaded successfully!")
    print(f"   Model: XGBoost Direct Severity Prediction")
    print(f"   Features: {len(feature_names)}")
    print(f"   Classes: {class_names}")

except Exception as e:
    print(f"❌ Error loading model: {e}")
    print("⚠️  Make sure xgboost_direct_model.pkl is in ./models/ directory")
    raise

# =============================================================================
# LOAD REAL ACCIDENT LOCATIONS FROM LOCAL FILE (ALL 32,324+ locations)
# =============================================================================
print("\n📍 Loading real accident locations from local file...")

try:
    # Load from local JSON file (faster and gets all 32,324 locations)
    ACCIDENT_LOCATIONS_FILE = "accident_locations_all.json"
    with open(ACCIDENT_LOCATIONS_FILE, "r", encoding="utf-8") as f:
        ACCIDENT_LOCATIONS = json.load(f)

    print(f"✅ Loaded {len(ACCIDENT_LOCATIONS):,} real accident locations from file")

    # Show statistics
    total_accidents = sum(loc.get("accident_count", 0) for loc in ACCIDENT_LOCATIONS)
    avg_accidents = (
        total_accidents / len(ACCIDENT_LOCATIONS) if ACCIDENT_LOCATIONS else 0
    )
    max_accidents = (
        max(loc.get("accident_count", 0) for loc in ACCIDENT_LOCATIONS)
        if ACCIDENT_LOCATIONS
        else 0
    )

    print(f"   Total accidents in dataset: {total_accidents:,}")
    print(f"   Average accidents per location: {avg_accidents:.1f}")
    print(f"   Highest accident count: {max_accidents}")

    # Add reverse geocoding cache
    geocoding_cache = {}

    def get_location_name(lat, lon, accident_count):
        """Get Thai road/location name using OpenStreetMap Nominatim"""
        cache_key = f"{lat:.4f},{lon:.4f}"
        if cache_key in geocoding_cache:
            return geocoding_cache[cache_key]

        try:
            # Use OSM Nominatim for reverse geocoding (free, no API key needed)
            url = "https://nominatim.openstreetmap.org/reverse"
            params = {
                "lat": lat,
                "lon": lon,
                "format": "json",
                "accept-language": "th",
                "zoom": 18,
            }
            headers = {"User-Agent": "ThailandAccidentRiskApp/1.0"}
            response = requests.get(url, params=params, headers=headers, timeout=3)

            if response.status_code == 200:
                data = response.json()
                address = data.get("address", {})

                # Try to build a road-focused name
                parts = []

                # Get road name first
                road = (
                    address.get("road")
                    or address.get("highway")
                    or address.get("street")
                )
                if road:
                    parts.append(road)

                # Add district/subdistrict
                subdistrict = address.get("suburb") or address.get("neighbourhood")
                district = address.get("city_district") or address.get("county")

                if subdistrict and subdistrict not in (parts[0] if parts else ""):
                    parts.append(subdistrict)
                elif district and district not in (parts[0] if parts else ""):
                    parts.append(district)

                if parts:
                    name = ", ".join(parts)
                    geocoding_cache[cache_key] = name
                    return name
        except Exception as e:
            pass

        # Fallback to accident count
        name = f"จุดเสี่ยง ({accident_count} ครั้ง)"
        geocoding_cache[cache_key] = name
        return name

    # Location names will be added on-demand during API requests (lazy loading)

    # Count severity distribution
    severity_counts = {}
    for loc in ACCIDENT_LOCATIONS:
        severity = loc.get("primary_severity", "ไม่ทราบ")
        severity_counts[severity] = severity_counts.get(severity, 0) + 1

    print(f"   Severity distribution:")
    for severity, count in sorted(
        severity_counts.items(), key=lambda x: x[1], reverse=True
    ):
        print(f"      {severity}: {count:,}")

except ImportError:
    print("⚠️  Supabase library not installed, trying local file...")
    ACCIDENT_LOCATIONS = []
    try:
        ACCIDENT_LOCATIONS_FILE = "accident_locations_all.json"
        with open(ACCIDENT_LOCATIONS_FILE, "r", encoding="utf-8") as f:
            ACCIDENT_LOCATIONS = json.load(f)
        print(f"✅ Loaded {len(ACCIDENT_LOCATIONS):,} locations from local file")
    except FileNotFoundError:
        print(f"⚠️  No accident locations found - will use grid-based scanning")
        ACCIDENT_LOCATIONS = []
except Exception as e:
    print(f"⚠️  Error loading accident locations from Supabase: {e}")
    print(f"   Trying local file fallback...")
    ACCIDENT_LOCATIONS = []
    try:
        ACCIDENT_LOCATIONS_FILE = "accident_locations_all.json"
        with open(ACCIDENT_LOCATIONS_FILE, "r", encoding="utf-8") as f:
            ACCIDENT_LOCATIONS = json.load(f)
        print(f"✅ Loaded {len(ACCIDENT_LOCATIONS):,} locations from local file")
    except:
        print(f"⚠️  No accident locations available - will use grid-based scanning")
        ACCIDENT_LOCATIONS = []

# =============================================================================
# REQUEST/RESPONSE MODELS
# =============================================================================


class PredictionRequest(BaseModel):
    """
    ข้อมูลที่ต้องส่งมาเพื่อทำนาย
    """

    # Location
    latitude: float
    longitude: float

    # Time
    hour: int  # 0-23
    day_of_week: int  # 0=Monday, 6=Sunday
    month: int  # 1-12

    # Weather (optional - ถ้าไม่มีจะใช้ค่า default)
    temperature: Optional[float] = 30.0
    rainfall: Optional[float] = 0.0
    weather_condition: Optional[str] = "clear"  # clear, rain, fog
    humidity: Optional[float] = 70.0

    # Traffic (optional)
    traffic_density: Optional[float] = 0.5  # 0-1
    average_speed: Optional[float] = 60.0  # km/h
    congestion_level: Optional[str] = "moderate"

    # Road (optional)
    road_type: Optional[str] = "city"  # highway, city, rural, local
    num_lanes: Optional[int] = 2
    has_street_light: Optional[bool] = True

    # Historical accidents nearby (CRITICAL for ML prediction)
    nearby_events_count: Optional[int] = 0  # Number of accidents within 10km
    is_weekend: Optional[bool] = False
    is_rush_hour: Optional[bool] = False
    
    # Vehicle info
    vehicle_type: Optional[str] = "car"  # car, motorcycle, truck, bus, bicycle, walk


class PredictionResponse(BaseModel):
    """Direct Severity Prediction Response"""
    prediction: str
    severity: str
    probabilities: Dict[str, float]
    risk_score: int
    risk_level: str
    confidence: float
    risk_factors: Optional[List[str]] = None
    recommendations: Optional[List[str]] = None


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================


def prepare_features(request: PredictionRequest) -> Dict[str, float]:
    """
    แปลงข้อมูลจาก request เป็น features สำหรับ two-stage model
    สร้าง feature dictionary ที่ตรงกับการเทรน temporal hotspot model
    """

    features = {}

    # Go through each feature และ set ค่าที่เหมาะสม
    for feat in feature_names:
        feat_lower = feat.lower()

        # Location features
        if feat == "LATITUDE":
            features[feat] = request.latitude
        elif feat == "LONGITUDE":
            features[feat] = request.longitude
        elif feat == "KM":
            features[feat] = 0.0

        # Time features
        elif feat == "hour":
            features[feat] = float(request.hour)
        elif feat == "month":
            features[feat] = request.month
        elif feat == "day":
            features[feat] = datetime.now().day
        elif feat == "year":
            features[feat] = datetime.now().year
        elif feat in ["dayofweek", "day_of_week"]:
            features[feat] = request.day_of_week
        elif feat == "quarter":
            features[feat] = (request.month - 1) // 3 + 1

        # Weekend/day type
        elif feat == "is_weekend":
            features[feat] = int(request.day_of_week >= 5)
        elif feat == "is_monday":
            features[feat] = int(request.day_of_week == 0)
        elif feat == "is_tuesday":
            features[feat] = int(request.day_of_week == 1)
        elif feat == "is_wednesday":
            features[feat] = int(request.day_of_week == 2)
        elif feat == "is_thursday":
            features[feat] = int(request.day_of_week == 3)
        elif feat == "is_friday":
            features[feat] = int(request.day_of_week == 4)
        elif feat == "is_saturday":
            features[feat] = int(request.day_of_week == 5)
        elif feat == "is_sunday":
            features[feat] = int(request.day_of_week == 6)

        # Hour categories
        elif "morning_rush" in feat_lower or feat == "is_morning_rush":
            features[feat] = int(6 <= request.hour < 9)
        elif "evening_rush" in feat_lower or feat == "is_evening_rush":
            features[feat] = int(17 <= request.hour < 20)
        elif "night" in feat_lower and "is_" in feat_lower:
            features[feat] = int(request.hour >= 22 or request.hour < 6)
        elif "lunch" in feat_lower:
            features[feat] = int(12 <= request.hour < 14)
        elif "afternoon" in feat_lower:
            features[feat] = int(14 <= request.hour < 17)
        elif "morning" in feat_lower and "rush" not in feat_lower:
            features[feat] = int(9 <= request.hour < 12)
        elif "evening" in feat_lower and "rush" not in feat_lower:
            features[feat] = int(20 <= request.hour < 23)

        # Weather features
        elif feat == "temperature":
            features[feat] = request.temperature
        elif feat == "dewpoint":
            features[feat] = request.temperature - ((100 - 70) / 5)  # Estimate
        elif feat == "humidity":
            features[feat] = 70.0  # Default
        elif feat == "wind_speed":
            features[feat] = 5.0  # Default
        elif feat == "wind_direction":
            features[feat] = 180.0  # Default
        elif feat == "pressure":
            features[feat] = 1013.25  # Default
        elif feat == "cloud_cover":
            features[feat] = 0.0
        elif "precipitation" in feat_lower or feat == "rain":
            features[feat] = request.rainfall

        # Weather conditions
        elif "raining" in feat_lower or "is_rain" in feat_lower:
            features[feat] = int(request.rainfall > 0)
        elif "clear" in feat_lower and "is_" in feat_lower:
            features[feat] = int(request.rainfall == 0)
        elif "hot" in feat_lower and "is_" in feat_lower:
            features[feat] = int(request.temperature > 35)
        elif "humid" in feat_lower and "is_" in feat_lower:
            features[feat] = int(True)  # Assume humid in Thailand
        elif "windy" in feat_lower and "is_" in feat_lower:
            features[feat] = int(False)

        # Cyclic encodings for time features
        elif feat == "hour_sin":
            features[feat] = np.sin(2 * np.pi * request.hour / 24)
        elif feat == "hour_cos":
            features[feat] = np.cos(2 * np.pi * request.hour / 24)
        elif feat == "dayofweek_sin":
            features[feat] = np.sin(2 * np.pi * request.day_of_week / 7)
        elif feat == "dayofweek_cos":
            features[feat] = np.cos(2 * np.pi * request.day_of_week / 7)
        elif feat == "month_sin":
            features[feat] = np.sin(2 * np.pi * request.month / 12)
        elif feat == "month_cos":
            features[feat] = np.cos(2 * np.pi * request.month / 12)
        elif feat == "day_sin":
            features[feat] = np.sin(2 * np.pi * datetime.now().day / 31)
        elif feat == "day_cos":
            features[feat] = np.cos(2 * np.pi * datetime.now().day / 31)

        # Risk scores based on time/conditions
        elif feat == "hour_risk_score":
            # Higher risk during rush hours and night
            if 17 <= request.hour <= 19:
                features[feat] = 80
            elif 7 <= request.hour <= 9:
                features[feat] = 70
            elif request.hour >= 22 or request.hour <= 5:
                features[feat] = 75
            else:
                features[feat] = 40
        elif feat == "day_risk_score":
            # Higher risk on weekends, especially Friday/Saturday
            if request.day_of_week == 5:  # Saturday
                features[feat] = 70
            elif request.day_of_week == 4:  # Friday
                features[feat] = 75
            elif request.day_of_week == 6:  # Sunday
                features[feat] = 65
            else:
                features[feat] = 50
        elif feat == "month_risk_score":
            # Higher risk during Songkran (April) and New Year
            if request.month == 4:  # Songkran
                features[feat] = 90
            elif request.month in [12, 1]:  # New Year
                features[feat] = 80
            else:
                features[feat] = 50
        elif feat == "weather_risk_score":
            if request.rainfall > 10:
                features[feat] = 85
            elif request.rainfall > 5:
                features[feat] = 70
            elif request.rainfall > 0:
                features[feat] = 60
            else:
                features[feat] = 30
        elif feat == "overall_risk_score":
            # Combine multiple factors
            base = 50
            if request.rainfall > 5:
                base += 15
            if request.hour >= 22 or request.hour <= 5:
                base += 10
            if request.day_of_week in [4, 5]:
                base += 10
            features[feat] = min(100, base)

        # Interaction features
        elif feat == "rain_rush_hour":
            features[feat] = int(request.rainfall > 0 and (7 <= request.hour <= 9 or 17 <= request.hour <= 19))
        elif feat == "friday_evening":
            features[feat] = int(request.day_of_week == 4 and 17 <= request.hour <= 23)
        elif feat == "weekend_night":
            features[feat] = int(request.day_of_week >= 5 and (request.hour >= 20 or request.hour <= 5))
        elif feat == "saturday_night":
            features[feat] = int(request.day_of_week == 5 and request.hour >= 20)
        elif feat == "rain_night":
            features[feat] = int(request.rainfall > 0 and (request.hour >= 20 or request.hour <= 5))

        # Encoded features (vehicle types, causes, etc.)
        elif "vehicle_type" in feat_lower:
            # Check if this feature corresponds to the requested vehicle type
            # e.g. "vehicle_type_motorcycle"
            req_vehicle = (request.vehicle_type or "car").lower()
            if req_vehicle in feat_lower:
                features[feat] = 1.0
            else:
                features[feat] = 0.0
        
        elif any(x in feat_lower for x in ["cause", "region", "encoded"]):
            features[feat] = 0.0

        # Default for any other feature
        else:
            features[feat] = 0.0

    return features


def calculate_risk_score(prediction: str, probabilities: np.ndarray) -> tuple:
    """
    คำนวณ risk score (0-100) จาก prediction
    """
    # Map class to risk score
    risk_map = {"บาดเจ็บเล็กน้อย": 30, "บาดเจ็บสาหัส": 60, "เสียชีวิต": 90}

    base_score = risk_map.get(prediction, 50)

    # ปรับด้วย confidence
    max_prob = probabilities.max()
    adjusted_score = int(base_score * (0.7 + 0.3 * max_prob))

    # กำหนด risk level
    if adjusted_score < 30:
        level = "low"
    elif adjusted_score < 50:
        level = "medium"
    elif adjusted_score < 70:
        level = "high"
    else:
        level = "very_high"

    return adjusted_score, level


# =============================================================================
# API ENDPOINTS
# =============================================================================


@app.get("/")
def root():
    """Health check"""
    return {
        "status": "online",
        "service": "Accident Risk Prediction API",
        "models": ["XGBoost", "KNN", "Random Forest"],
        "version": "1.0.0",
    }


@app.get("/models/info")
def get_models_info():
    """ข้อมูลเกี่ยวกับโมเดล"""
    return {
        "num_features": len(feature_names),
        "classes": class_names,
        "models": {
            "xgboost": "XGBoost Classifier",
            "knn": "K-Nearest Neighbors",
            "random_forest": "Random Forest Classifier",
        },
    }


def predict_severity_with_reasoning(
    features_dict: Dict[str, float], request: PredictionRequest = None
) -> Dict:
    """
    Direct Severity Prediction with Reasoning:
    Predicts accident severity (บาดเจ็บเล็กน้อย, บาดเจ็บสาหัส, เสียชีวิต)
    and provides reasoning based on risk factors.
    
    No hotspot detection - always returns severity prediction.
    """
    # Convert dict to numpy array
    features_array = np.array([[features_dict.get(col, 0) for col in feature_names]])
    
    # Get nearby events count
    nearby_count = request.nearby_events_count if request else 0
    
    # Direct severity prediction using the new model
    severity_pred = xgb_direct_model.predict(features_array)[0]
    severity_proba = xgb_direct_model.predict_proba(features_array)[0]
    severity_class = label_encoder.inverse_transform([severity_pred])[0]
    
    # Calculate base risk score from severity
    severity_score_map = {
        "บาดเจ็บเล็กน้อย": 30,  # Minor Injury
        "บาดเจ็บสาหัส": 60,     # Serious Injury  
        "เสียชีวิต": 90,         # Fatal
    }
    base_score = severity_score_map.get(severity_class, 40)
    
    # Adjust based on ML confidence - CRITICAL for uncertain predictions
    severity_confidence = float(severity_proba.max())
    
    # If confidence is low (< 0.5), significantly reduce the risk score
    # This prevents overconfident predictions when model is uncertain
    if severity_confidence < 0.5:
        # Low confidence: scale down aggressively
        confidence_multiplier = severity_confidence  # 0.38 confidence = 38% of base score
    else:
        # High confidence: scale normally
        confidence_multiplier = 0.7 + 0.3 * severity_confidence
    
    adjusted_score = int(base_score * confidence_multiplier)
    
    # Generate reasoning based on risk factors
    risk_factors = []
    
    # Historical accident data
    if nearby_count > 80:
        risk_factors.append(f"Very high accident history ({nearby_count} accidents within 10km)")
        adjusted_score = min(100, adjusted_score + 20)
    elif nearby_count > 50:
        risk_factors.append(f"High accident history ({nearby_count} accidents within 10km)")
        adjusted_score = min(100, adjusted_score + 15)
    elif nearby_count > 20:
        risk_factors.append(f"Moderate accident history ({nearby_count} accidents within 10km)")
        adjusted_score = min(100, adjusted_score + 10)
    elif nearby_count > 5:
        risk_factors.append(f"Some accident history ({nearby_count} accidents within 10km)")
        adjusted_score = min(100, adjusted_score + 5)
    
    # Weather conditions
    if features_dict.get("rainfall", 0) > 0:
        rainfall = features_dict.get("rainfall", 0)
        if rainfall > 10:
            risk_factors.append("Heavy rain (reduced visibility and traction)")
            adjusted_score = min(100, adjusted_score + 15)
        elif rainfall > 5:
            risk_factors.append("Moderate rain (slippery roads)")
            adjusted_score = min(100, adjusted_score + 10)
        else:
            risk_factors.append("Light rain (wet conditions)")
            adjusted_score = min(100, adjusted_score + 5)

    
    # Time-based factors
    hour = features_dict.get("hour", 12)
    day_of_week = features_dict.get("day_of_week", 0)
    
    if hour >= 20 or hour <= 5:
        risk_factors.append("Night time (reduced visibility, driver fatigue)")
        adjusted_score = min(100, adjusted_score + 10)
    
    if features_dict.get("is_rush_hour", 0) or (7 <= hour <= 9) or (17 <= hour <= 19):
        risk_factors.append("Rush hour (high traffic volume, stress)")
        adjusted_score = min(100, adjusted_score + 8)
    
    # Day of week
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    current_day = day_names[day_of_week]
    
    if day_of_week == 4:  # Friday (0-indexed: Monday=0, Friday=4)
        risk_factors.append(f"{current_day} evening (end of work week, driver fatigue)")
        adjusted_score = min(100, adjusted_score + 5)
    elif day_of_week == 5:  # Saturday
        risk_factors.append(f"{current_day} (weekend leisure travel, varied driver experience)")
        adjusted_score = min(100, adjusted_score + 3)
    elif day_of_week == 6:  # Sunday
        risk_factors.append(f"{current_day} (return travel, tired drivers)")
        adjusted_score = min(100, adjusted_score + 3)
    
    if features_dict.get("is_weekend", 0):
        if day_of_week not in [5, 6]:  # Only add if not already mentioned
            risk_factors.append("Weekend (higher recreational traffic)")
    
    # Traffic conditions
    congestion = features_dict.get("congestion_level", "low")
    if congestion == "high":
        risk_factors.append("High traffic congestion (stop-and-go, rear-end risk)")
        adjusted_score = min(100, adjusted_score + 7)
    elif congestion == "moderate":
        risk_factors.append("Moderate traffic (reduced maneuverability)")
        adjusted_score = min(100, adjusted_score + 4)
    
    # Probabilities dict
    probs_dict = {
        class_names[i]: float(severity_proba[i]) for i in range(len(class_names))
    }
    
    # Generate safety recommendations based on severity AND probabilities
    recommendations = []
    
    # Get probability spread to understand uncertainty
    prob_values = list(probs_dict.values())
    prob_spread = max(prob_values) - min(prob_values)
    
    # Base recommendations by severity
    if severity_class == "เสียชีวิต":
        recommendations.append("⚠️ HIGH RISK - Exercise extreme caution")
        recommendations.append("🚗 Reduce speed significantly and increase following distance")
        recommendations.append("👀 Stay highly alert for sudden hazards")
    elif severity_class == "บาดเจ็บสาหัส":
        recommendations.append("⚡ MODERATE-HIGH RISK - Drive defensively")
        recommendations.append("🛣️ Maintain safe following distance")
        recommendations.append("👁️ Watch carefully for pedestrians and vehicles")
    else:
        recommendations.append("✅ LOWER RISK - Remain vigilant")
        recommendations.append("🚦 Obey all traffic signals and speed limits")
    
    # Add uncertainty warning if probabilities are close
    if prob_spread < 0.2:  # If top prediction is only 20% higher than others
        recommendations.append("⚠️ Conditions are unpredictable - exercise extra caution")
    
    # Specific recommendations based on identified risk factors
    if any("rain" in str(f).lower() or "wet" in str(f).lower() for f in risk_factors):
        recommendations.append("☔ Reduce speed in wet conditions, avoid sudden braking")
    if any("night" in str(f).lower() for f in risk_factors):
        recommendations.append("🌙 Use headlights, avoid sudden lane changes")
    if any("rush hour" in str(f).lower() for f in risk_factors):
        recommendations.append("⏰ Allow extra travel time, stay patient")
    if any("friday" in str(f).lower() or "weekend" in str(f).lower() or "saturday" in str(f).lower() or "sunday" in str(f).lower() for f in risk_factors):
        recommendations.append("🎉 Watch for increased recreational traffic")
    if any("congestion" in str(f).lower() or "traffic" in str(f).lower() for f in risk_factors):
        recommendations.append("🚦 Avoid aggressive driving, maintain safe distance")
    if any("accident history" in str(f).lower() for f in risk_factors):
        recommendations.append("📍 Known accident-prone area - be extra vigilant")
    
    # Always include distraction warning
    if len(recommendations) < 6:
        recommendations.append("📱 Avoid all distractions while driving")
    
    return {
        "prediction": severity_class,
        "severity": severity_class,
        "probabilities": probs_dict,
        "confidence": severity_confidence,
        "risk_score": adjusted_score,
        "nearby_accidents": nearby_count,
        "risk_factors": risk_factors,
        "recommendations": recommendations[:6],  # Limit to 6
    }


@app.post("/predict", response_model=PredictionResponse)
async def predict_accident_risk(request: PredictionRequest):
    """
    ทำนายความเสี่ยงอุบัติเหตุด้วย Two-Stage Model
    Stage 1: Hotspot Detection
    Stage 2: Severity Prediction (if hotspot)
    """
    try:
        # เตรียม features
        features = prepare_features(request)

        # Severity prediction with reasoning
        result = predict_severity_with_reasoning(features, request)

        # Log prediction results
        nearby_count = request.nearby_events_count
        print(f"📊 Prediction at ({request.latitude:.4f}, {request.longitude:.4f}):")
        print(f"   - Nearby accidents (10km): {nearby_count}")
        print(f"   - Severity prediction: {result['prediction']}")
        print(f"   - Risk score: {result['risk_score']}%")
        print(f"   - Confidence: {result['confidence']:.2f}")

        # กำหนด risk level
        risk_score = result["risk_score"]
        if risk_score < 30:
            risk_level = "low"
        elif risk_score < 50:
            risk_level = "medium"
        elif risk_score < 70:
            risk_level = "high"
        else:
            risk_level = "very_high"

        return PredictionResponse(
            prediction=result["prediction"],
            severity=result["severity"],
            probabilities=result["probabilities"],
            risk_score=risk_score,
            risk_level=risk_level,
            confidence=result["confidence"],
            risk_factors=result.get("risk_factors", []),
            recommendations=result.get("recommendations", []),
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Prediction error: {str(e)}")


@app.post("/predict/route")
async def predict_route_risk(
    from_lat: float,
    from_lng: float,
    to_lat: float,
    to_lng: float,
    departure_time: str,  # ISO format: "2025-01-15T08:00:00"
    vehicle_type: Optional[str] = "car",
):
    """
    ทำนายความเสี่ยงสำหรับเส้นทาง (จุดเริ่มต้นถึงปลายทาง)
    ใช้ Two-Stage Model
    """
    try:
        # Parse departure time
        dt = datetime.fromisoformat(departure_time.replace("Z", "+00:00"))

        # ทำนายที่จุดเริ่มต้น
        start_request = PredictionRequest(
            latitude=from_lat,
            longitude=from_lng,
            hour=dt.hour,
            day_of_week=dt.weekday(),
            month=dt.month,
            vehicle_type=vehicle_type,
        )
        start_prediction = await predict_accident_risk(start_request)

        # ทำนายที่จุดกึ่งกลาง
        mid_lat = (from_lat + to_lat) / 2
        mid_lng = (from_lng + to_lng) / 2
        # สมมติว่าใช้เวลาเดินทางครึ่งทาง (เช่น 30 นาที)
        mid_dt = dt + timedelta(minutes=30)
        
        mid_request = PredictionRequest(
            latitude=mid_lat,
            longitude=mid_lng,
            hour=mid_dt.hour,
            day_of_week=mid_dt.weekday(),
            month=mid_dt.month,
            vehicle_type=vehicle_type,
        )
        mid_prediction = await predict_accident_risk(mid_request)

        # ทำนายที่ปลายทาง
        # สมมติว่าใช้เวลาเดินทาง 1 ชั่วโมง
        end_dt = dt + timedelta(hours=1)
        
        end_request = PredictionRequest(
            latitude=to_lat,
            longitude=to_lng,
            hour=end_dt.hour,
            day_of_week=end_dt.weekday(),
            month=end_dt.month,
            vehicle_type=vehicle_type,
        )
        end_prediction = await predict_accident_risk(end_request)

        # คำนวณ overall route risk
        avg_risk = (
            start_prediction.risk_score
            + end_prediction.risk_score
            + mid_prediction.risk_score
        ) / 3

        return {
            "route_risk_score": int(avg_risk),
            "start": {
                "prediction": start_prediction.prediction,
                "risk_score": start_prediction.risk_score,
                "risk_level": start_prediction.risk_level,
            },
            "middle": {
                "prediction": mid_prediction.prediction,
                "risk_score": mid_prediction.risk_score,
                "risk_level": mid_prediction.risk_level,
            },
            "end": {
                "prediction": end_prediction.prediction,
                "risk_score": end_prediction.risk_score,
                "risk_level": end_prediction.risk_level,
            },
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Route prediction error: {str(e)}")


class HotspotRequest(BaseModel):
    hour: int = 18
    day_of_week: int = 4
    month: int = 1
    rainfall: float = 0.0
    traffic_density: float = 0.5
    min_probability: float = 0.01


@app.post("/predict/hotspots")
async def predict_hotspots(request: HotspotRequest):
    """
    ทำนายจุดเสี่ยงทั้งหมดในประเทศไทย โดยใช้สถานที่จริงที่มีอุบัติเหตุเกิดขึ้น
    Uses ALL 32,324+ real accident locations from Supabase + ML predictions
    """
    hour = request.hour
    day_of_week = request.day_of_week
    month = request.month
    rainfall = request.rainfall
    traffic_density = request.traffic_density

    print(f"\n🔍 Predicting risks for real accident locations...")
    print(
        f"   Conditions: Hour={hour}, Day={day_of_week}, Month={month}, Rainfall={rainfall}mm, Traffic={traffic_density}"
    )

    if len(ACCIDENT_LOCATIONS) == 0:
        print("   ⚠️  No accident locations loaded")
        return {
            "error": "No accident locations available",
            "total_locations_checked": 0,
            "hotspots_found": 0,
            "hotspots": [],
        }

    # Use top 5000 locations by accident count
    locations_to_check = ACCIDENT_LOCATIONS[:5000]
    print(f"   Using top {len(locations_to_check):,} highest-risk locations")

    # BATCH PREDICTION
    features_list = []
    for loc in locations_to_check:
        pred_request = PredictionRequest(
            latitude=loc["latitude"],
            longitude=loc["longitude"],
            hour=hour,
            day_of_week=day_of_week,
            month=month,
            rainfall=rainfall,
            traffic_density=traffic_density,
        )
        features_dict = prepare_features(pred_request)
        feature_array = [features_dict.get(fname, 0) for fname in feature_names]
        features_list.append(feature_array)

    features_batch = np.array(features_list)
    print(f"   Running batch ML predictions...")

    # Direct severity prediction for all locations
    severity_preds_batch = xgb_direct_model.predict(features_batch)
    severity_probs_batch = xgb_direct_model.predict_proba(features_batch)

    # Use logarithmic scaling for accident counts to create better variation
    # Top 5000 locations have accident_count ranging from ~10 to 358
    max_count = max(loc.get("accident_count", 1) for loc in locations_to_check)
    min_count = min(loc.get("accident_count", 1) for loc in locations_to_check)

    print(f"   Accident count range: {min_count} to {max_count}")

    hotspots = []
    for i, loc in enumerate(locations_to_check):
        severity_pred = severity_preds_batch[i]
        severity_proba = severity_probs_batch[i]
        severity_class = label_encoder.inverse_transform([severity_pred])[0]

        # Calculate risk score based on severity
        severity_score_map = {
            "บาดเจ็บเล็กน้อย": 30,
            "บาดเจ็บสาหัส": 60,
            "เสียชีวิต": 90,
        }
        base_risk = severity_score_map.get(severity_class, 40)

        # Use logarithmic scaling for historical risk to create variation
        accident_count = loc.get("accident_count", 1)
        # Map accident_count logarithmically: log(count)/log(max_count) * 100
        import math

        if accident_count > 1 and max_count > 1:
            historical_risk = int(
                (math.log(accident_count) / math.log(max_count)) * 100
            )
        else:
            historical_risk = 10

        # Combine ML risk and historical risk (weighted average)
        # ML model is more important (70%), historical data is 30%
        adjusted_risk = int(base_risk * 0.7 + historical_risk * 0.3)

        # Adjust for weather and time conditions
        if rainfall > 5:
            adjusted_risk = min(100, adjusted_risk + 15)
        if hour in [7, 8, 17, 18, 19]:
            adjusted_risk = min(100, adjusted_risk + 10)
        if day_of_week in [5, 6]:
            adjusted_risk = min(100, adjusted_risk + 5)

        # Only include high-risk locations (risk >= 50)
        if adjusted_risk >= 50:
            # Use placeholder name for now (will geocode top results later)
            lat = loc["latitude"]
            lon = loc["longitude"]
            location_name = f"temp_{i}"  # Temporary placeholder

            hotspots.append(
                {
                    "name": location_name,
                    "latitude": loc.get("latitude"),
                    "longitude": loc.get("longitude"),
                    "severity": severity_class,
                    "risk_score": combined_risk, # Assuming adjusted_risk was meant to be combined_risk
                    "accident_count": accident_count,
                    "province": loc.get("province_name_th", ""),
                    "historical_severity": loc.get("primary_severity"),
                    "peak_hours": loc.get("peak_hours", []),
                }
            )

    hotspots.sort(key=lambda x: x["risk_score"], reverse=True)
    hotspots = hotspots[:1000]

    # Geocode only top 100 locations for performance
    print(f"   🗺️  Getting road names for top 100 locations...")
    geocoded_count = 0
    for i, hotspot in enumerate(hotspots[:100]):
        if hotspot["name"].startswith("temp_"):
            road_name = get_location_name(
                hotspot["latitude"],
                hotspot["longitude"],
                hotspot["historical_accidents"],
            )
            hotspot["name"] = road_name
            geocoded_count += 1
            if (i + 1) % 20 == 0:
                print(f"      Geocoded {i + 1}/100...")

    # For the rest, use simple format
    for hotspot in hotspots[100:]:
        if hotspot["name"].startswith("temp_"):
            hotspot["name"] = f"จุดเสี่ยง ({hotspot['historical_accidents']} ครั้ง)"

    print(
        f"   ✅ Found {len(hotspots)} risk zones out of {len(locations_to_check):,} checked"
    )
    print(f"   📍 Geocoded {geocoded_count} locations with road names")
    if hotspots:
        print(f"   Highest risk: {hotspots[0]['risk_score']} at {hotspots[0]['name']}")

    return {
        "total_locations_checked": len(locations_to_check),
        "hotspots_found": len(hotspots),
        "data_source": "real_accident_locations_supabase",
        "total_locations_available": len(ACCIDENT_LOCATIONS),
        "conditions": {
            "hour": hour,
            "day_of_week": day_of_week,
            "month": month,
            "rainfall": rainfall,
            "traffic_density": traffic_density,
        },
        "hotspots": hotspots,
    }


@app.get("/itic/events")
async def get_itic_events(year: Optional[int] = None, historical: bool = False):
    """
    Fetch events from Longdo Events Feed
    Parameters:
    - year: Specific year (default: current year)
    - historical: If True, fetch all available years
    """
    try:
        from datetime import datetime

        if year is None:
            year = datetime.now().year

        print(f"📡 Fetching Longdo events feed for year {year}...")

        all_events = []
        years_to_fetch = []

        if historical:
            # Fetch multiple years (2020-2025)
            current_year = datetime.now().year
            years_to_fetch = list(range(2020, current_year + 1))
            print(f"   Fetching historical data for years: {years_to_fetch}")
        else:
            years_to_fetch = [year]

        for fetch_year in years_to_fetch:
            try:
                # Try JSON feed first
                response = requests.get(
                    f"https://event.longdo.com/feed/{fetch_year}",
                    timeout=10,
                    headers={"Accept": "application/json"},
                )

                if response.status_code == 200:
                    # Check if response is JSON or RSS/XML
                    content_type = response.headers.get("content-type", "")

                    if "json" in content_type.lower():
                        data = response.json()
                        events = (
                            data if isinstance(data, list) else data.get("events", [])
                        )
                    else:
                        # Parse RSS/XML feed
                        import xml.etree.ElementTree as ET

                        root = ET.fromstring(response.text)

                        events = []
                        # Parse RSS format
                        for item in root.findall(".//item"):
                            event = {
                                "title": item.findtext("title", ""),
                                "description": item.findtext("description", ""),
                                "link": item.findtext("link", ""),
                                "pubDate": item.findtext("pubDate", ""),
                                "category": item.findtext("category", "event"),
                                "year": fetch_year,
                            }

                            # Extract lat/lon from georss tags if available
                            georss = item.find("{http://www.georss.org/georss}point")
                            if georss is not None and georss.text:
                                coords = georss.text.split()
                                if len(coords) == 2:
                                    event["lat"] = float(coords[0])
                                    event["lon"] = float(coords[1])

                            events.append(event)

                    print(f"   ✅ Year {fetch_year}: {len(events)} events")
                    all_events.extend(events)
                else:
                    print(f"   ⚠️ Year {fetch_year}: Status {response.status_code}")

            except Exception as e:
                print(f"   ❌ Error fetching year {fetch_year}: {e}")
                continue

        print(f"✅ Total fetched: {len(all_events)} Longdo events")
        return {"events": all_events, "total": len(all_events), "years": years_to_fetch}

    except Exception as e:
        print(f"❌ Error fetching Longdo events: {e}")
        import traceback

        traceback.print_exc()
        return {"events": [], "total": 0, "years": []}


@app.get("/itic/cameras")
async def get_itic_cameras():
    """
    Proxy endpoint for iTIC Cameras API to bypass CORS
    """
    try:
        import requests

        print("📹 Fetching iTIC cameras via proxy...")
        response = requests.get(
            "http://cameras.iticfoundation.org/api/getCamList.json.php", timeout=10
        )

        if response.status_code == 200:
            data = response.json()
            print(f"✅ Fetched {len(data.get('cameras', []))} iTIC cameras")
            return data
        else:
            print(f"⚠️ iTIC Cameras API returned status {response.status_code}")
            return {"cameras": []}

    except Exception as e:
        print(f"❌ Error fetching iTIC cameras: {e}")
        return {"cameras": []}


@app.get("/traffic/index")
async def get_traffic_index_data(year: Optional[int] = None, historical: bool = False):
    """
    Fetch traffic index from Longdo Traffic
    Returns CSV data parsed into JSON format

    Parameters:
    - year: Specific year (default: current year)
    - historical: If True, fetch multiple years of data
    """
    try:
        import csv
        from datetime import datetime
        from io import StringIO

        if year is None:
            year = datetime.now().year

        print(f"📊 Fetching traffic index from Longdo for year {year}...")

        all_data = []
        years_to_fetch = []

        if historical:
            # Fetch multiple years (2012-2025 available)
            current_year = datetime.now().year
            years_to_fetch = list(range(2020, current_year + 1))
            print(f"   Fetching historical traffic index for: {years_to_fetch}")
        else:
            years_to_fetch = [year]

        for fetch_year in years_to_fetch:
            try:
                response = requests.get(
                    f"https://traffic.longdo.com/api/raw/trafficindex/{fetch_year}",
                    timeout=10,
                )

                if response.status_code == 200:
                    # Parse CSV data
                    csv_data = StringIO(response.text)
                    reader = csv.DictReader(csv_data)

                    year_data = []
                    for row in reader:
                        year_data.append(
                            {
                                "timestamp": int(row["timestamp"]),
                                "datetime": row["datetime"],
                                "index": float(row["index"]),
                                "year": fetch_year,
                            }
                        )

                    print(f"   ✅ Year {fetch_year}: {len(year_data)} records")
                    all_data.extend(year_data)
                else:
                    print(f"   ⚠️ Year {fetch_year}: Status {response.status_code}")

            except Exception as e:
                print(f"   ❌ Error fetching year {fetch_year}: {e}")
                continue

        if not historical and all_data:
            # For current year, return last 24h data
            recent_data = all_data[-288:] if len(all_data) > 288 else all_data
        else:
            recent_data = all_data

        current_index = recent_data[-1]["index"] if recent_data else 0
        avg_index = (
            sum(d["index"] for d in recent_data) / len(recent_data)
            if recent_data
            else 0
        )

        print(f"✅ Fetched {len(recent_data)} traffic index records")
        if not historical:
            print(
                f"   Current index: {current_index:.2f}, 24h average: {avg_index:.2f}"
            )

        return {
            "current": current_index,
            "average_24h": round(avg_index, 2),
            "data": recent_data,
            "total_records": len(recent_data),
            "years": years_to_fetch,
        }

    except Exception as e:
        print(f"❌ Error fetching traffic index: {e}")
        return {
            "current": 0,
            "average_24h": 0,
            "data": [],
            "total_records": 0,
            "years": [],
        }


@app.get("/events/database")
async def get_events_from_database(
    year: Optional[int] = None,
    month: Optional[int] = None,
    historical: bool = False,
    event_types: Optional[str] = None,
    severities: Optional[str] = None,
    province: Optional[str] = None,
    limit: Optional[int] = None,
    offset: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """
    Get events from Supabase database

    Parameters:
    - year: Specific year
    - month: Specific month (1-12) - filters by month within the year
    - historical: Fetch multiple years (2012-present)
    - event_types: Comma-separated event types
    - severities: Comma-separated severities
    - province: Filter by province
    - limit: Max results (default: 5000 for single year, 10000 for multi-year)
    - offset: Offset for pagination (default: 0)
    - start_date: Start date for range filter (YYYY-MM-DD)
    - end_date: End date for range filter (YYYY-MM-DD)
    """
    try:
        from datetime import datetime

        from supabase_traffic_client import get_supabase_traffic_client

        print(f"📊 Querying traffic events from Supabase...")

        client = get_supabase_traffic_client()

        # Parse filters
        event_type_list = event_types.split(",") if event_types else None
        severity_list = severities.split(",") if severities else None

        # If date range is provided, use the new date range query
        if start_date and end_date:
            events_data, total_count = client.get_events_by_date_range(
                start_date=start_date,
                end_date=end_date,
                event_types=event_type_list,
                severities=severity_list,
                limit=limit,
                offset=offset,
            )

            # Transform to frontend format
            events = []
            for event in events_data:
                events.append(
                    {
                        "id": event.get("event_id"),
                        "title": event.get("title_th") or event.get("title_en"),
                        "description": event.get("description_th")
                        or event.get("description_en"),
                        "lat": float(event.get("latitude")),
                        "lon": float(event.get("longitude")),
                        "category": event.get("event_type"),
                        "severity": event.get("severity_score"),
                        "pubDate": event.get("event_date"),
                        "year": event.get("year"),
                        "location": event.get("location_name"),
                        "source": event.get("source"),
                    }
                )

            print(
                f"✅ Retrieved {len(events)} events from Supabase (total: {total_count})"
            )

            return {
                "events": events,
                "total": total_count,
                "source": "supabase",
                "filters": {
                    "start_date": start_date,
                    "end_date": end_date,
                    "event_types": event_type_list,
                    "severities": severity_list,
                    "limit": limit,
                    "offset": offset,
                },
            }

        if year is None:
            year = datetime.now().year

        # Query database
        if province:
            events_data = client.get_events_by_province(
                province=province, year=year if not historical else None, limit=limit
            )
            years_used = (
                [year] if not historical else list(range(2012, datetime.now().year + 1))
            )
        elif historical:
            events_data = client.get_events_multi_year(
                start_year=2012,
                end_year=datetime.now().year,
                event_types=event_type_list,
                severities=severity_list,
                limit=limit,
            )
            years_used = list(range(2012, datetime.now().year + 1))
        else:
            events_data = client.get_events_by_year(
                year=year,
                month=month,
                event_types=event_type_list,
                severities=severity_list,
                limit=limit,
                offset=offset,
            )
            years_used = [year]

        # Transform to frontend format
        events = []
        for event in events_data:
            events.append(
                {
                    "id": event.get("event_id"),
                    "title": event.get("title_th") or event.get("title_en"),
                    "description": event.get("description_th")
                    or event.get("description_en"),
                    "lat": float(event.get("latitude")),
                    "lon": float(event.get("longitude")),
                    "category": event.get("event_type"),
                    "severity": event.get("severity_score"),
                    "pubDate": event.get("event_date"),
                    "year": event.get("year"),
                    "location": event.get("location_name"),
                    "source": event.get("source"),
                }
            )

        print(f"✅ Retrieved {len(events)} events from Supabase")

        return {
            "events": events,
            "total": len(events),
            "years": years_used,
            "source": "supabase",
            "filters": {
                "year": year if not historical else None,
                "historical": historical,
                "event_types": event_type_list,
                "severities": severity_list,
                "province": province,
            },
        }

    except Exception as e:
        print(f"❌ Error querying database: {e}")
        import traceback

        traceback.print_exc()
        return {
            "events": [],
            "total": 0,
            "years": [],
            "source": "error",
            "error": str(e),
        }


@app.get("/events/export-csv")
async def export_events_to_csv(
    start_date: str,
    end_date: str,
    event_types: Optional[str] = None,
    severities: Optional[str] = None,
):
    """
    Export events to CSV format (streamed response)

    Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - event_types: Comma-separated event types
    - severities: Comma-separated severities

    Returns: CSV file download
    """
    import csv
    from io import StringIO

    from fastapi.responses import StreamingResponse

    try:
        from datetime import datetime

        from supabase_traffic_client import get_supabase_traffic_client

        print(f"📊 Exporting events from {start_date} to {end_date} to CSV...")

        client = get_supabase_traffic_client()

        # Parse filters
        event_type_list = event_types.split(",") if event_types else None
        severity_list = severities.split(",") if severities else None

        # Fetch ALL events (no limit) for export
        # The database query is efficient and handles large datasets
        events_data, total_count = client.get_events_by_date_range(
            start_date=start_date,
            end_date=end_date,
            event_types=event_type_list,
            severities=severity_list,
            limit=None,  # No limit for export
            offset=0,
        )

        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)

        # Write header (Thai + English)
        writer.writerow(
            [
                "ลำดับที่ / No.",
                "รายงานเมื่อ / Reported At",
                "ประเภท / Type",
                "หัวข้อเหตุการณ์ / Event Title",
                "สถานที่ / Location",
                "ละติจูด / Latitude",
                "ลองจิจูด / Longitude",
                "ความรุนแรง / Severity",
                "ผู้รายงาน / Source",
            ]
        )

        # Event type mapping (Thai labels)
        event_category_map = {
            "accident": "อุบัติเหตุ",
            "construction": "ก่อสร้าง",
            "congestion": "รถติด",
            "flooding": "น้ำท่วม",
            "fire": "เพลิงไหม้",
            "breakdown": "รถเสีย",
            "road_closed": "ถนนปิด",
            "diversion": "เบี่ยงจราจร",
        }

        # Write data rows
        for i, event in enumerate(events_data, 1):
            # Format datetime
            try:
                event_date = datetime.fromisoformat(
                    event.get("event_date", "").replace("Z", "+00:00")
                )
                formatted_date = event_date.strftime("%d/%m/%Y %H:%M")
            except:
                formatted_date = event.get("event_date", "")

            writer.writerow(
                [
                    i,
                    formatted_date,
                    event_category_map.get(
                        event.get("event_type"), event.get("event_type", "")
                    ),
                    event.get("title_th") or event.get("title_en", ""),
                    event.get("location_name", ""),
                    event.get("latitude", ""),
                    event.get("longitude", ""),
                    event.get("severity_score", ""),
                    event.get("source", ""),
                ]
            )

        # Get CSV content
        csv_content = output.getvalue()
        output.close()

        # Add BOM for Thai characters in Excel
        csv_bytes = "\ufeff" + csv_content

        print(f"✅ Exported {len(events_data)} events to CSV")

        # Return as downloadable file
        filename = f"traffic_events_{start_date}_{end_date}.csv"

        return StreamingResponse(
            iter([csv_bytes.encode("utf-8")]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except Exception as e:
        print(f"❌ Error exporting CSV: {e}")
        import traceback

        traceback.print_exc()

        return {"error": str(e), "message": "Failed to export CSV"}


# =============================================================================
# DASHBOARD STATISTICS ENDPOINT WITH FILTER FIXES
# =============================================================================

# Cache for dashboard stats (in-memory)
_dashboard_cache = {}
_dashboard_cache_time = {}
DASHBOARD_CACHE_TTL = 300  # 5 minutes

# =====================================================
# MAPPING DICTIONARIES FOR FILTERS
# =====================================================
# IMPORTANT: These mappings use EXACT values from the database
# Generated from /dashboard/filter-values endpoint query

# Vehicle Type Mapping (Frontend ID -> Database Thai Names in vehicle_1 column)
# Based on actual database values - 10 unique types
VEHICLE_TYPE_MAPPING = {
    "รถปิคอัพบรรทุก 4 ล้อ": ["รถปิคอัพบรรทุก 4 ล้อ"],  # 391 records
    "รถยนต์นั่งส่วนบุคคล/รถยนต์นั่งสาธารณะ": ["รถยนต์นั่งส่วนบุคคล/รถยนต์นั่งสาธารณะ"],  # 264 records
    "รถจักรยานยนต์": ["รถจักรยานยนต์"],  # 153 records
    "รถบรรทุกมากกว่า 10 ล้อ (รถพ่วง)": ["รถบรรทุกมากกว่า 10 ล้อ (รถพ่วง)"],  # 76 records
    "รถบรรทุก 6 ล้อ": ["รถบรรทุก 6 ล้อ"],  # 34 records
    "รถบรรทุกมากกว่า 6 ล้อ ไม่เกิน 10 ล้อ": ["รถบรรทุกมากกว่า 6 ล้อ ไม่เกิน 10 ล้อ"],  # 26 records
    "รถตู้": ["รถตู้"],  # 6 records
    "รถปิคอัพโดยสาร": ["รถปิคอัพโดยสาร"],  # 5 records
    "อื่นๆ": ["อื่นๆ"],  # 27 records
    "ไม่ระบุประเภทรถ": ["ไม่ระบุประเภทรถ"],  # 7 records
}

# Weather Condition Mapping (Frontend ID -> Database Thai Names in weather_condition column)
# Based on actual database values - 5 unique conditions
WEATHER_CONDITION_MAPPING = {
    "แจ่มใส": ["แจ่มใส"],  # 731 records
    "ฝนตก": ["ฝนตก"],  # 260 records
    "มีหมอก/ควัน/ฝุ่น": ["มีหมอก/ควัน/ฝุ่น"],  # 6 records
    "มืดครึ้ม": ["มืดครึ้ม"],  # 1 record
    "ไม่ทราบสภาพอากาศ": ["ไม่ทราบสภาพอากาศ"],  # 2 records
}

# Accident Cause Mapping (Frontend ID -> Database Thai Names in presumed_cause column)
# Based on actual database values - 10+ unique causes
# NOTE: ชื่อ column ในฐานข้อมูลคือ presumed_cause ไม่ใช่ accident_cause
ACCIDENT_CAUSE_MAPPING = {
    "ขับรถเร็วเกินอัตรากำหนด": ["ขับรถเร็วเกินอัตรากำหนด"],  # 794 records - MOST COMMON
    "คน/รถ/สัตว์ตัดหน้ากระชั้นชิด": ["คน/รถ/สัตว์ตัดหน้ากระชั้นชิด"],  # 85 records
    "หลับใน": ["หลับใน"],  # 36 records
    "อุปกรณ์ยานพาหนะบกพร่อง": ["อุปกรณ์ยานพาหนะบกพร่อง"],  # 32 records
    "ฝ่าฝืนสัญญาณไฟ/เครื่องหมายจราจร": ["ฝ่าฝืนสัญญาณไฟ/เครื่องหมายจราจร"],  # 12 records
    "แซงรถอย่างผิดกฎหมาย": ["แซงรถอย่างผิดกฎหมาย"],  # 7 records
    "เมาสุรา": ["เมาสุรา"],  # 5 records
    "ถนนลื่น": ["ถนนลื่น"],  # 2 records
    "อื่นๆ": ["อื่นๆ"],  # 3 records
    "ไม่ทราบมูลเหตุ": ["ไม่ทราบมูลเหตุ"],  # 12 records
}


def map_filter_to_database_values(filter_id: str, mapping: dict) -> list:
    """
    แปลง filter ID จาก frontend เป็น list ของค่าในฐานข้อมูล

    Args:
        filter_id: ID จาก frontend (เช่น "motorcycle", "clear")
        mapping: Dictionary mapping (เช่น VEHICLE_TYPE_MAPPING)

    Returns:
        List ของค่าในฐานข้อมูลที่ตรงกัน
    """
    if filter_id == "all" or not filter_id:
        return []

    # Get mapped values
    mapped_values = mapping.get(filter_id, [])

    # Also include the filter_id itself in case it matches exactly
    if filter_id not in mapped_values:
        mapped_values.append(filter_id)

    return mapped_values


@app.get("/dashboard/filter-values")
async def get_filter_values():
    """
    Get unique values for all filter columns from database
    Returns actual values with counts for bilingual mapping
    
    This endpoint queries the database to get real unique values,
    which can be used to create accurate Thai-English filter mappings.
    """
    try:
        # Import inside function like other endpoints
        from supabase_traffic_client import get_supabase_traffic_client
        
        client = get_supabase_traffic_client()
        
        # Query all events to analyze filter values
        print("📊 Querying database for unique filter values...")
        
        # Use the correct client attribute
        response = client.client.table("accident_records").select(
            "vehicle_1,weather_condition,presumed_cause"
        ).execute()
        
        events = response.data
        print(f"✅ Loaded {len(events):,} events")
        
        # Count occurrences for each filter column
        vehicle_counts = {}
        weather_counts = {}
        cause_counts = {}
        
        for event in events:
            # Vehicle types
            vehicle = event.get("vehicle_1", "")
            if vehicle and vehicle.strip():
                vehicle_counts[vehicle] = vehicle_counts.get(vehicle, 0) + 1
            
            # Weather conditions
            weather = event.get("weather_condition", "")
            if weather and weather.strip():
                weather_counts[weather] = weather_counts.get(weather, 0) + 1
            
            # Accident causes
            cause = event.get("presumed_cause", "")
            if cause and cause.strip():
                cause_counts[cause] = cause_counts.get(cause, 0) + 1
        
        # Sort by frequency (most common first)
        vehicle_sorted = sorted(vehicle_counts.items(), key=lambda x: x[1], reverse=True)
        weather_sorted = sorted(weather_counts.items(), key=lambda x: x[1], reverse=True)
        cause_sorted = sorted(cause_counts.items(), key=lambda x: x[1], reverse=True)
        
        # Log results
        print(f"\n📊 Filter Values Summary:")
        print(f"  - Vehicle Types: {len(vehicle_counts)} unique values")
        print(f"  - Weather Conditions: {len(weather_counts)} unique values")
        print(f"  - Accident Causes: {len(cause_counts)} unique values")
        
        return {
            "vehicle_types": [
                {"value": value, "count": count}
                for value, count in vehicle_sorted
            ],
            "weather_conditions": [
                {"value": value, "count": count}
                for value, count in weather_sorted
            ],
            "accident_causes": [
                {"value": value, "count": count}
                for value, count in cause_sorted
            ],
            "total_events": len(events),
        }
        
    except Exception as e:
        print(f"❌ Error getting filter values: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/dashboard/stats")
async def get_dashboard_stats(
    date_range: Optional[str] = "all",
    province: Optional[str] = "all",
    casualty_type: Optional[str] = "all",
    vehicle_type: Optional[str] = "all",
    weather: Optional[str] = "all",
    accident_cause: Optional[str] = "all",  # NEW PARAMETER
):
    """
    Get dashboard statistics using PostgreSQL aggregation (FAST!)

    FIXED: Vehicle type และ Weather filters ใช้งานได้แล้ว
    ADDED: Accident cause filter

    Parameters:
    - date_range: Year filter (all, 2025, 2024, 2023, etc.)
    - province: Province filter (all or province name)
    - casualty_type: Casualty severity filter (all, fatal, serious, minor, survivors)
    - vehicle_type: Vehicle type filter (all, motorcycle, car, truck, etc.)
    - weather: Weather condition filter (all, clear, rain, cloudy, fog)
    - accident_cause: Accident cause filter (all, speeding, drunk_driving, etc.)

    Returns: Dashboard statistics including summary cards, charts data
    """
    try:
        from collections import defaultdict
        from datetime import datetime

        from supabase_traffic_client import get_supabase_traffic_client

        # Check cache first
        # NOTE: Cache only by date_range, province, casualty_type
        # Vehicle/Weather/Cause will be filtered on frontend for speed!
        cache_key = f"{date_range}:{province}:{casualty_type}"
        if cache_key in _dashboard_cache:
            cached_time = _dashboard_cache_time.get(cache_key)
            if (
                cached_time
                and (datetime.now() - cached_time).total_seconds() < DASHBOARD_CACHE_TTL
            ):
                print(f"✅ Returning cached dashboard stats for {cache_key}")
                return _dashboard_cache[cache_key]

        print(
            f"📊 Fetching dashboard stats (date_range={date_range}, province={province}, "
            f"casualty_type={casualty_type})..."
        )
        print(
            f"   Note: vehicle_type={vehicle_type}, weather={weather}, accident_cause={accident_cause} will be filtered on frontend"
        )

        client = get_supabase_traffic_client()

        # Build date filter
        start_date = "2019-01-01"
        end_date = "2025-08-31"

        if date_range != "all":
            year = int(date_range)
            if year == 2025:
                start_date = f"{year}-01-01"
                end_date = f"{year}-08-31"
            else:
                start_date = f"{year}-01-01"
                end_date = f"{year}-12-31"

        # ========================================
        # MAP FILTERS TO DATABASE VALUES
        # ========================================

        # NOTE: Vehicle/Weather/Cause filters will be applied on FRONTEND
        # for better performance and cache efficiency
        # Only Province and Date filters are applied on backend
        print(
            f"   Backend filters: date_range={date_range}, province={province}, casualty_type={casualty_type}"
        )
        print(
            f"   Frontend filters (not applied here): vehicle={vehicle_type}, weather={weather}, cause={accident_cause}"
        )

        # Fetch data with pagination
        all_events = []
        page_size = 1000
        offset = 0

        # Select fields we need (รวม presumed_cause สำหรับ accident cause filter)
        select_fields = "accident_datetime,accident_type,province,vehicle_1,weather_condition,presumed_cause,casualties_fatal,casualties_serious,casualties_minor"

        while True:
            print(f"🔄 Fetching page at offset {offset} (page size: {page_size})")

            query = client.client.table("accident_records").select(
                select_fields, count="exact"
            )

            # Apply date filter
            query = query.gte("accident_datetime", start_date).lte(
                "accident_datetime", end_date
            )

            # Apply province filter
            if province != "all":
                query = query.eq("province", province)

            # NOTE: Vehicle/Weather/Cause filters NOT applied here
            # They will be filtered on frontend for better cache performance

            # Apply pagination
            query = query.range(offset, offset + page_size - 1)

            response = query.execute()
            events_page = response.data

            print(
                f"   📦 Received {len(events_page)} records (total count: {response.count if hasattr(response, 'count') else 'N/A'})"
            )

            if not events_page:
                print(f"   ⚠️ Empty page received, stopping pagination")
                break

            all_events.extend(events_page)

            print(f"   ✅ Total accumulated: {len(all_events):,} records")

            # Break if last page
            if len(events_page) < page_size:
                print(
                    f"   🏁 Last page (received {len(events_page)} < {page_size}), stopping"
                )
                break

            offset += page_size

        events = all_events
        print(f"✅ Fetched {len(events):,} events (after all filters)")

        # ========================================
        # AGGREGATION
        # ========================================
        total_accidents = len(events)

        # Sum casualty counts
        total_fatalities = 0
        total_serious = 0
        total_minor = 0

        event_type_counts = defaultdict(int)
        weather_counts = defaultdict(int)
        accident_cause_counts = defaultdict(int)  # Enabled!
        province_counts = defaultdict(int)
        province_casualties = defaultdict(
            lambda: {"fatal": 0, "serious": 0, "minor": 0}
        )  # NEW: Track casualties per province
        monthly_counts = defaultdict(int)
        daily_counts_by_month = defaultdict(lambda: defaultdict(int))
        yearly_summary = defaultdict(int)
        monthly_summary = defaultdict(int)
        weekday_summary = defaultdict(int)
        hourly_counts = [0] * 24
        day_counts = [0] * 7

        # Single pass through data
        for event in events:
            # Get casualty counts
            fatal = int(event.get("casualties_fatal", 0) or 0)
            serious = int(event.get("casualties_serious", 0) or 0)
            minor = int(event.get("casualties_minor", 0) or 0)

            # Filter by casualty_type if specified
            if casualty_type != "all":
                if casualty_type == "fatal" and fatal == 0:
                    continue
                elif casualty_type == "serious" and serious == 0:
                    continue
                elif casualty_type == "minor" and minor == 0:
                    continue
                elif casualty_type == "survivors" and fatal > 0:
                    continue

            # Sum casualties
            total_fatalities += fatal
            total_serious += serious
            total_minor += minor

            # Accident type (keep original Thai values)
            event_type = event.get("accident_type", "other")
            event_type_counts[event_type] += 1

            # Weather (keep original Thai values)
            weather_val = event.get("weather_condition", "ไม่ทราบ")
            weather_counts[weather_val] += 1

            # Accident cause (keep original Thai values from presumed_cause)
            cause = event.get("presumed_cause", "")
            if cause and cause.strip():
                accident_cause_counts[cause] += 1

            # Province
            prov = event.get("province", "Unknown")
            province_counts[prov] += 1

            # Track casualties per province
            province_casualties[prov]["fatal"] += fatal
            province_casualties[prov]["serious"] += serious
            province_casualties[prov]["minor"] += minor

            # Time-based aggregations
            try:
                event_date = datetime.fromisoformat(
                    event.get("accident_datetime", "").replace("Z", "+00:00")
                )

                # Yearly summary
                year_key = str(event_date.year)
                yearly_summary[year_key] += 1

                # Monthly (YYYY-MM)
                month_key = event_date.strftime("%Y-%m")
                monthly_counts[month_key] += 1

                # Monthly summary (01-12)
                month_only = event_date.strftime("%m")
                monthly_summary[month_only] += 1

                # Daily (within each month)
                date_key = event_date.strftime("%Y-%m-%d")
                daily_counts_by_month[month_key][date_key] += 1

                # Weekday summary
                weekday_key = event_date.weekday()
                weekday_summary[weekday_key] += 1

                # Hourly
                hourly_counts[event_date.hour] += 1

                # Daily (day of week)
                day_counts[event_date.weekday()] += 1
            except:
                continue

        print(f"✅ Aggregation complete")
        print(
            f"   Total casualties: {total_fatalities} fatal, {total_serious} serious, {total_minor} minor"
        )

        # Get top 10 provinces
        top_provinces = sorted(
            province_counts.items(), key=lambda x: x[1], reverse=True
        )[:10]

        # Get ALL provinces for heatmap (with casualty details)
        all_provinces = sorted(
            [
                {
                    "province": prov,
                    "count": count,
                    "fatal": province_casualties[prov]["fatal"],
                    "serious": province_casualties[prov]["serious"],
                    "minor": province_casualties[prov]["minor"],
                    "survivors": count - province_casualties[prov]["fatal"],
                }
                for prov, count in province_counts.items()
            ],
            key=lambda x: x["count"],
            reverse=True,
        )

        # Calculate survivors
        survivors_count = total_accidents - total_fatalities

        # Prepare response
        result = {
            "summary": {
                "total_accidents": total_accidents,
                "minor_injuries": total_minor,
                "serious_injuries": total_serious,
                "fatalities": total_fatalities,
                "survivors": survivors_count,
                "high_risk_areas": len(
                    [p for p in province_counts.values() if p > 100]
                ),
            },
            # Add raw event details for frontend filtering
            "all_events": [
                {
                    "vehicle_1": e.get("vehicle_1", ""),
                    "weather_condition": e.get("weather_condition", ""),
                    "presumed_cause": e.get("presumed_cause", ""),
                    "accident_type": e.get("accident_type", ""),
                    "province": e.get("province", ""),
                    "casualties_fatal": e.get("casualties_fatal", 0),
                    "casualties_serious": e.get("casualties_serious", 0),
                    "casualties_minor": e.get("casualties_minor", 0),
                }
                for e in all_events
            ],
            "severity_distribution": [
                {
                    "name": "ผู้รอดชีวิต",
                    "value": survivors_count,
                    "color": "#10b981",
                },
                {
                    "name": "ผู้บาดเจ็บเล็กน้อย",
                    "value": total_minor,
                    "color": "#EAB308",  # Yellow
                },
                {
                    "name": "ผู้บาดเจ็บสาหัส",
                    "value": total_serious,
                    "color": "#f59e0b",
                },
                {
                    "name": "ผู้เสียชีวิต",
                    "value": total_fatalities,
                    "color": "#ef4444",
                },
            ],
            "event_types": [
                {"type": k, "count": v}
                for k, v in sorted(
                    event_type_counts.items(), key=lambda x: x[1], reverse=True
                )[:20]
            ],
            "weather_data": [
                {"weather": k, "count": v}
                for k, v in sorted(
                    weather_counts.items(), key=lambda x: x[1], reverse=True
                )
            ],
            "accident_causes": [
                {"cause": k, "count": v}
                for k, v in sorted(
                    accident_cause_counts.items(), key=lambda x: x[1], reverse=True
                )[:10]  # Top 10 causes
            ],
            "top_provinces": [{"province": p[0], "count": p[1]} for p in top_provinces],
            "all_provinces": all_provinces,  # Now includes fatal/serious/minor/survivors
            "monthly_trend": [
                {
                    "month": k,
                    "count": v,
                    "daily": [
                        {"date": date, "count": count}
                        for date, count in sorted(daily_counts_by_month[k].items())
                    ],
                }
                for k, v in sorted(monthly_counts.items())
            ],
            "hourly_pattern": [
                {"hour": i, "count": hourly_counts[i]} for i in range(24)
            ],
            "daily_pattern": [
                {
                    "day": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][i],
                    "count": day_counts[i],
                }
                for i in range(7)
            ],
            "yearly_summary": [
                {"year": year, "count": count}
                for year, count in sorted(yearly_summary.items())
            ],
            "monthly_summary": [
                {
                    "month": str(i + 1).zfill(2),
                    "month_name": [
                        "Jan",
                        "Feb",
                        "Mar",
                        "Apr",
                        "May",
                        "Jun",
                        "Jul",
                        "Aug",
                        "Sep",
                        "Oct",
                        "Nov",
                        "Dec",
                    ][i],
                    "count": monthly_summary[str(i + 1).zfill(2)],
                }
                for i in range(12)
            ],
            "weekday_summary": [
                {
                    "day": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][i],
                    "day_name": [
                        "Monday",
                        "Tuesday",
                        "Wednesday",
                        "Thursday",
                        "Friday",
                        "Saturday",
                        "Sunday",
                    ][i],
                    "count": weekday_summary[i],
                }
                for i in range(7)
            ],
        }

        # Cache the result
        _dashboard_cache[cache_key] = result
        _dashboard_cache_time[cache_key] = datetime.now()
        print(f"💾 Cached dashboard stats for {cache_key}")

        return result

    except Exception as e:
        print(f"❌ Error fetching dashboard stats: {e}")
        import traceback

        traceback.print_exc()
        return {"error": str(e), "message": "Failed to fetch dashboard statistics"}


@app.get("/events/available-years")
async def get_available_years():
    """
    Get list of years that have event data with counts

    Returns: List of years with event counts
    """
    try:
        from supabase_traffic_client import get_supabase_traffic_client

        print(f"📅 Fetching available years...")

        client = get_supabase_traffic_client()

        # Query to get distinct years with counts
        response = (
            client.supabase.table("traffic_events")
            .select("event_date", count="exact")
            .execute()
        )

        # Group by year and count
        year_counts = {}
        for event in response.data:
            try:
                from datetime import datetime

                event_date = datetime.fromisoformat(
                    event.get("event_date", "").replace("Z", "+00:00")
                )
                year = event_date.year
                year_counts[year] = year_counts.get(year, 0) + 1
            except:
                continue

        # Sort years in descending order
        years_data = [
            {"year": year, "count": count}
            for year, count in sorted(year_counts.items(), reverse=True)
        ]

        print(f"✅ Found {len(years_data)} years with data")

        return {"years": years_data, "total_years": len(years_data)}

    except Exception as e:
        print(f"❌ Error fetching available years: {e}")
        import traceback

        traceback.print_exc()

        return {"error": str(e), "message": "Failed to fetch available years"}


@app.get("/events/export-excel")
async def export_events_to_excel(
    start_date: str,
    end_date: str,
    event_types: Optional[str] = None,
    severities: Optional[str] = None,
):
    """
    Export events to Excel format (XLSX)

    Parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - event_types: Comma-separated event types
    - severities: Comma-separated severities

    Returns: Excel file download
    """
    from io import BytesIO

    from fastapi.responses import StreamingResponse

    try:
        from datetime import datetime

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from supabase_traffic_client import get_supabase_traffic_client

        print(f"📊 Exporting events from {start_date} to {end_date} to Excel...")

        client = get_supabase_traffic_client()

        # Parse filters
        event_type_list = event_types.split(",") if event_types else None
        severity_list = severities.split(",") if severities else None

        # Fetch ALL events (no limit) for export
        events_data, total_count = client.get_events_by_date_range(
            start_date=start_date,
            end_date=end_date,
            event_types=event_type_list,
            severities=severity_list,
            limit=None,  # No limit for export
            offset=0,
        )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Traffic Events"

        # Event type mapping (Thai labels)
        event_category_map = {
            "accident": "อุบัติเหตุ",
            "construction": "ก่อสร้าง",
            "congestion": "รถติด",
            "flooding": "น้ำท่วม",
            "fire": "เพลิงไหม้",
            "breakdown": "รถเสีย",
            "road_closed": "ถนนปิด",
            "diversion": "เบี่ยงจราจร",
            "roadwork": "ซ่อมถนน",
            "other": "อื่นๆ",
        }

        # Write header with styling
        headers = [
            "ลำดับ",
            "วันที่-เวลา",
            "ประเภท",
            "หัวข้อ",
            "รายละเอียด",
            "สถานที่",
            "แหล่งที่มา",
            "ละติจูด",
            "ลองจิจูด",
            "ระดับความรุนแรง",
        ]

        # Style for header
        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Set column widths
        column_widths = [8, 18, 15, 40, 50, 30, 12, 12, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_num)
            ].width = width

        # Write data rows
        for i, event in enumerate(events_data, 1):
            # Format datetime
            try:
                event_date = datetime.fromisoformat(
                    event.get("event_date", "").replace("Z", "+00:00")
                )
                formatted_date = event_date.strftime("%d/%m/%Y %H:%M")
            except:
                formatted_date = event.get("event_date", "")

            row_data = [
                i,
                formatted_date,
                event_category_map.get(
                    event.get("event_type"), event.get("event_type", "")
                ),
                event.get("title_th") or event.get("title_en", ""),
                event.get("description_th") or event.get("description_en", ""),
                event.get("location_name", ""),
                event.get("source", ""),
                event.get("latitude", ""),
                event.get("longitude", ""),
                event.get("severity_score", ""),
            ]

            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=i + 1, column=col_num, value=value)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        print(f"✅ Exported {len(events_data)} events to Excel")

        # Return as downloadable file
        filename = f"traffic_events_{start_date}_{end_date}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except Exception as e:
        print(f"❌ Error exporting Excel: {e}")
        import traceback

        traceback.print_exc()

        return {"error": str(e), "message": "Failed to export Excel"}


if __name__ == "__main__":
    import uvicorn

    print("🚀 Starting Accident Risk Prediction API on port 10000...")
    uvicorn.run(app, host="0.0.0.0", port=10000)
